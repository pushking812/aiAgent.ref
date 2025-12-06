#!/usr/bin/env python3
"""
Оптимизированный скрипт для создания Excel-отчета по МЕЖМОДУЛЬНЫМ связям Python-проекта.
"""

import os
import ast
import sys
from pathlib import Path
from collections import defaultdict, deque
from typing import Dict, List, Set, Tuple, Optional, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse
import time

class OptimizedDependencyReporter:
    def __init__(self, root_path: str, exclude_dirs: List[str] = None):
        self.root_path = Path(root_path).absolute()
        self.exclude_dirs = exclude_dirs or []
        self.exclude_dirs.extend(['venv', '.venv', '__pycache__', '.git', '.idea', '.pytest_cache'])
        
        # Индексы
        self.module_index: Dict[str, Path] = {}
        self.inter_module_relations: List[Dict] = []
        self.connection_matrix: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        
        # Оптимизированные кэши
        self.element_cache: Dict[str, Dict[str, List]] = {}
        self.quick_lookup: Dict[str, List[Dict]] = defaultdict(list)
        
        # Статистика
        self.stats = {
            'total_modules': 0,
            'inter_module_relations': 0,
            'files_analyzed': 0,
            'analysis_time': 0
        }
    
    def analyze_project(self) -> None:
        """Анализирует проект и собирает только межмодульные связи."""
        start_time = time.time()
        print(f"🔍 Анализ проекта: {self.root_path}")
        
        # 1. Быстрая индексация модулей
        self._quick_index_modules()
        
        # 2. Однопроходный анализ каждого модуля
        for module_name, file_path in self.module_index.items():
            self._fast_analyze_module(module_name, file_path)
        
        self.stats['analysis_time'] = time.time() - start_time
        
        print(f"✅ Анализ завершен за {self.stats['analysis_time']:.2f} сек.")
        print(f"   Всего модулей: {self.stats['total_modules']}")
        print(f"   Межмодульных связей: {self.stats['inter_module_relations']}")
    
    def _quick_index_modules(self) -> None:
        """Быстрая индексация Python-модулей в проекте."""
        for py_file in self.root_path.rglob("*.py"):
            # Быстрая проверка исключений
            if any(excluded in str(py_file) for excluded in self.exclude_dirs):
                continue
            
            rel_path = py_file.relative_to(self.root_path)
            
            if py_file.name == '__init__.py':
                module_name = str(rel_path.parent).replace(os.sep, '.')
            else:
                module_name = str(rel_path.with_suffix('')).replace(os.sep, '.')
            
            self.module_index[module_name] = py_file
        
        self.stats['total_modules'] = len(self.module_index)
    
    def _fast_analyze_module(self, module_name: str, file_path: Path) -> None:
        """Быстрый анализ одного модуля."""
        self.stats['files_analyzed'] += 1
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                lines = content.split('\n')
            
            tree = ast.parse(content)
            
            # Сначала строим быстрый кэш элементов этого модуля
            module_elements = self._extract_module_elements(module_name, tree, lines)
            self.element_cache[module_name] = module_elements
            
            # Добавляем в быстрый поисковый индекс
            for element in module_elements['all_elements']:
                self.quick_lookup[element['name']].append(element)
            
            # Анализируем связи
            self._analyze_links_in_module(module_name, tree, lines)
            
        except Exception as e:
            print(f"⚠️  Ошибка при анализе {file_path}: {e}")
    
    def _extract_module_elements(self, module_name: str, tree: ast.AST, lines: List[str]) -> Dict:
        """Извлекает все элементы модуля за один проход."""
        elements = {
            'classes': [],
            'functions': [],
            'methods': [],
            'all_elements': []
        }
        
        # Стек для отслеживания контекста
        context_stack = []
        
        for node in ast.walk(tree):
            # Обновляем стек контекста
            if isinstance(node, (ast.ClassDef, ast.FunctionDef, ast.AsyncFunctionDef)):
                context_stack.append(node)
            
            # Извлекаем информацию о классе
            if isinstance(node, ast.ClassDef):
                class_info = {
                    'type': 'class',
                    'name': node.name,
                    'module': module_name,
                    'line': node.lineno,
                    'full_name': f"{module_name}.{node.name}",
                    'bases': [self._get_base_name(base) for base in node.bases]
                }
                elements['classes'].append(class_info)
                elements['all_elements'].append(class_info)
            
            # Извлекаем информацию о функции/методе
            elif isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
                # Проверяем контекст
                in_class = any(isinstance(ctx, ast.ClassDef) for ctx in context_stack)
                
                if in_class:
                    # Это метод
                    current_class = next((ctx for ctx in reversed(context_stack) 
                                        if isinstance(ctx, ast.ClassDef)), None)
                    
                    method_info = {
                        'type': 'method',
                        'name': node.name,
                        'module': module_name,
                        'class': current_class.name if current_class else '',
                        'line': node.lineno,
                        'full_name': f"{module_name}.{current_class.name}.{node.name}" if current_class else f"{module_name}.{node.name}",
                        'is_async': isinstance(node, ast.AsyncFunctionDef)
                    }
                    elements['methods'].append(method_info)
                    elements['all_elements'].append(method_info)
                else:
                    # Это функция уровня модуля
                    func_info = {
                        'type': 'function',
                        'name': node.name,
                        'module': module_name,
                        'line': node.lineno,
                        'full_name': f"{module_name}.{node.name}",
                        'is_async': isinstance(node, ast.AsyncFunctionDef)
                    }
                    elements['functions'].append(func_info)
                    elements['all_elements'].append(func_info)
            
            # Выходим из контекста
            if node in context_stack:
                context_stack.remove(node)
        
        return elements
    
    def _get_base_name(self, base_node: ast.AST) -> str:
        """Получает имя базового класса."""
        if isinstance(base_node, ast.Name):
            return base_node.id
        elif isinstance(base_node, ast.Attribute):
            return ast.unparse(base_node)
        return ''
    
    def _analyze_links_in_module(self, source_module: str, tree: ast.AST, lines: List[str]) -> None:
        """Анализирует все ссылки в модуле за один проход."""
        visitor = LinkVisitor(source_module, self.module_index, self.element_cache, 
                            self.quick_lookup, lines)
        visitor.visit(tree)
        
        # Добавляем найденные связи
        for relation in visitor.relations:
            if relation['target_module'] and relation['target_module'] != source_module:
                self._add_relation(relation)
    
    def _add_relation(self, relation: Dict) -> None:
        """Добавляет межмодульную связь."""
        # Форматируем информацию об источнике
        source_context = relation.get('source_context', {})
        
        # Определяем, что указать в столбце метода/функции
        source_method = ''
        source_function = ''
        
        if source_context.get('method'):
            source_method = source_context['method']
        elif source_context.get('function'):
            source_function = source_context['function']
        
        # Форматируем информацию о цели
        target_info = self._get_target_info(relation['target_module'], 
                                          relation.get('target_element', ''))
        
        # Создаем запись о связи
        record = {
            'source_folder': self._extract_folder(relation['source_module']),
            'source_module': relation['source_module'],
            'source_class': source_context.get('class', ''),
            'source_function': source_function,
            'source_method': source_method,
            'source_line': relation['source_line'],
            'source_line_text': relation.get('source_line_text', '')[:200],
            
            'target_folder': self._extract_folder(relation['target_module']),
            'target_module': relation['target_module'],
            'target_class': target_info.get('class', ''),
            'target_function': target_info.get('function', ''),
            'target_method': target_info.get('method', ''),
            'target_line': target_info.get('line', 0),
            'target_line_text': target_info.get('line_text', '')[:200],
            
            'relation_type': relation['relation_type']
        }
        
        self.inter_module_relations.append(record)
        self.stats['inter_module_relations'] += 1
        
        # Обновляем матрицу связей
        self.connection_matrix[relation['source_module']][relation['target_module']] += 1
    
    def _get_target_info(self, target_module: str, target_element: str) -> Dict:
        """Получает информацию о целевом элементе."""
        if not target_element or target_module not in self.element_cache:
            return {}
        
        # Ищем элемент в кэше целевого модуля
        for element in self.element_cache[target_module]['all_elements']:
            if element['name'] == target_element or element['full_name'].endswith(f".{target_element}"):
                return {
                    'class': element.get('class', ''),
                    'function': element['name'] if element['type'] == 'function' else '',
                    'method': element['name'] if element['type'] == 'method' else '',
                    'line': element['line'],
                    'line_text': f"Defined at line {element['line']}"
                }
        
        return {}
    
    def _extract_folder(self, module_name: str) -> str:
        """Извлекает папку из имени модуля."""
        if '.' not in module_name:
            return ''
        return '.'.join(module_name.split('.')[:-1])
    
    def create_excel_report(self, output_file: str = "inter_module_dependencies.xlsx") -> None:
        """Создает Excel-отчет с межмодульными зависимостями."""
        print(f"\n📊 Создание Excel-отчета...")
        
        if not self.inter_module_relations:
            print("⚠️  Не найдено межмодульных зависимостей для отчета")
            return
        
        # Создаем Excel файл
        wb = Workbook()
        
        # 1. Лист с детальными связями
        self._create_detailed_sheet(wb)
        
        # 2. Лист с матрицей связей
        self._create_matrix_sheet(wb)
        
        # 3. Лист со сводкой
        self._create_summary_sheet(wb)
        
        # Удаляем дефолтный лист
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Сохраняем файл
        wb.save(output_file)
        
        print(f"✅ Отчет сохранен: {output_file}")
        self._print_console_summary()
    
    def _create_detailed_sheet(self, wb: Workbook) -> None:
        """Создает лист с детальными связями."""
        ws = wb.create_sheet(title="Межмодульные связи")
        
        # Заголовки
        headers = [
            'Тип связи', 'Источник: Папка', 'Источник: Модуль', 
            'Источник: Класс', 'Источник: Функция', 'Источник: Метод',
            'Источник: Строка', 'Источник: Код строки',
            'Цель: Папка', 'Цель: Модуль', 'Цель: Класс',
            'Цель: Функция', 'Цель: Метод', 'Цель: Строка',
            'Цель: Код строки'
        ]
        
        # Добавляем заголовки
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Добавляем данные
        for row_idx, rel in enumerate(self.inter_module_relations, start=2):
            data = [
                rel['relation_type'],
                rel['source_folder'],
                rel['source_module'],
                rel['source_class'],
                rel['source_function'],
                rel['source_method'],
                rel['source_line'],
                rel['source_line_text'],
                rel['target_folder'],
                rel['target_module'],
                rel['target_class'],
                rel['target_function'],
                rel['target_method'],
                rel['target_line'],
                rel['target_line_text']
            ]
            
            fill_color = self._get_row_color(rel['relation_type'])
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = self._thin_border()
        
        # Настраиваем ширину столбцов
        widths = [15, 20, 30, 20, 20, 20, 10, 40, 20, 30, 20, 20, 20, 10, 40]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
    
    def _create_matrix_sheet(self, wb: Workbook) -> None:
        """Создает лист с матрицей связей."""
        ws = wb.create_sheet(title="Матрица связей")
        
        # Получаем все модули с связями
        all_modules = set()
        for src in self.connection_matrix:
            all_modules.add(src)
            for dst in self.connection_matrix[src]:
                all_modules.add(dst)
        
        all_modules = sorted(all_modules)
        
        if not all_modules:
            ws.cell(row=1, column=1, value="Нет межмодульных связей")
            return
        
        # Заголовок
        ws.cell(row=1, column=1, value="ШАХМАТНАЯ МАТРИЦА СВЯЗЕЙ МЕЖДУ МОДУЛЯМИ").font = Font(bold=True, size=14)
        ws.merge_cells(f'A1:{get_column_letter(len(all_modules) + 2)}1')
        
        # Заголовки столбцов
        ws.cell(row=2, column=1, value="Модуль (источник) →").font = Font(bold=True)
        for col_idx, module in enumerate(all_modules, start=2):
            cell = ws.cell(row=2, column=col_idx, value=module)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True, text_rotation=45)
        
        # Заголовки строк
        for row_idx, module in enumerate(all_modules, start=3):
            cell = ws.cell(row=row_idx, column=1, value=module)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="right")
        
        # Заполняем матрицу
        for row_idx, target in enumerate(all_modules, start=3):
            for col_idx, source in enumerate(all_modules, start=2):
                count = self.connection_matrix[source].get(target, 0)
                if count > 0:
                    cell = ws.cell(row=row_idx, column=col_idx, value=count)
                    # Цветовая градация
                    if count >= 10:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif count >= 5:
                        cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif count >= 2:
                        cell.fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value="")
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                cell.border = self._thin_border()
        
        # Добавляем итоги
        total_row = len(all_modules) + 3
        total_col = len(all_modules) + 2
        
        # Итоги по столбцам (исходящие)
        ws.cell(row=total_row, column=1, value="ВСЕГО исходящих").font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        for col_idx, source in enumerate(all_modules, start=2):
            total_out = sum(self.connection_matrix[source].values())
            cell = ws.cell(row=total_row, column=col_idx, value=total_out)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Итоги по строкам (входящие)
        ws.cell(row=2, column=total_col, value="ВСЕГО входящих").font = Font(bold=True)
        ws.cell(row=2, column=total_col).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        for row_idx, target in enumerate(all_modules, start=3):
            total_in = sum(self.connection_matrix[src].get(target, 0) for src in all_modules)
            cell = ws.cell(row=row_idx, column=total_col, value=total_in)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Общий итог
        ws.cell(row=total_row, column=total_col, 
                value=sum(sum(d.values()) for d in self.connection_matrix.values()))
        cell = ws.cell(row=total_row, column=total_col)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        # Настраиваем ширину
        ws.column_dimensions['A'].width = 40
        for i in range(2, total_col + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15
        
        ws.freeze_panes = 'C3'
    
    def _create_summary_sheet(self, wb: Workbook) -> None:
        """Создает лист со сводкой."""
        ws = wb.create_sheet(title="Сводка")
        
        # Заголовок
        ws.merge_cells('A1:E1')
        ws.cell(row=1, column=1, value="СВОДКА МЕЖМОДУЛЬНЫХ ЗАВИСИМОСТЕЙ").font = Font(bold=True, size=14)
        
        # Статистика
        row = 3
        ws.cell(row=row, column=1, value="Общая статистика").font = Font(bold=True, size=12)
        
        stats = [
            ("Проект", str(self.root_path)),
            ("Дата анализа", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Время анализа", f"{self.stats['analysis_time']:.2f} сек."),
            ("Всего модулей", self.stats['total_modules']),
            ("Межмодульных связей", self.stats['inter_module_relations']),
            ("Файлов проанализировано", self.stats['files_analyzed'])
        ]
        
        for i, (label, value) in enumerate(stats, start=row+1):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
        
        # Топ модулей
        row = len(stats) + row + 2
        
        # Подсчитываем активность модулей
        module_stats = {}
        all_modules = set()
        
        for src, targets in self.connection_matrix.items():
            all_modules.add(src)
            for dst in targets:
                all_modules.add(dst)
        
        for module in all_modules:
            outgoing = sum(self.connection_matrix[module].values())
            incoming = sum(self.connection_matrix[src].get(module, 0) for src in all_modules)
            module_stats[module] = {'outgoing': outgoing, 'incoming': incoming, 'total': outgoing + incoming}
        
        # Топ по исходящим
        top_outgoing = sorted(module_stats.items(), key=lambda x: x[1]['outgoing'], reverse=True)[:10]
        
        ws.cell(row=row, column=1, value="ТОП-10 модулей по исходящим связям").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:C{row}')
        
        for i, (module, stats) in enumerate(top_outgoing, start=row+1):
            ws.cell(row=i, column=1, value=module)
            ws.cell(row=i, column=2, value=stats['outgoing'])
        
        # Топ по входящим
        row = row + len(top_outgoing) + 2
        top_incoming = sorted(module_stats.items(), key=lambda x: x[1]['incoming'], reverse=True)[:10]
        
        ws.cell(row=row, column=1, value="ТОП-10 модулей по входящим связям").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:C{row}')
        
        for i, (module, stats) in enumerate(top_incoming, start=row+1):
            ws.cell(row=i, column=1, value=module)
            ws.cell(row=i, column=2, value=stats['incoming'])
        
        # Настраиваем ширину
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
    
    def _get_row_color(self, relation_type: str) -> str:
        """Возвращает цвет строки."""
        colors = {
            'import': 'E6F3FF',
            'import_from': 'D9F2E6', 
            'call': 'FFF2E6',
            'attribute': 'F2E6FF',
            'class_ref': 'FFE6E6'
        }
        return colors.get(relation_type, 'FFFFFF')
    
    def _thin_border(self) -> Border:
        """Создает тонкую границу."""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _print_console_summary(self) -> None:
        """Выводит сводку в консоль."""
        print(f"\n📋 СВОДКА ОТЧЕТА:")
        print("=" * 80)
        
        # Статистика по типам связей
        type_counts = defaultdict(int)
        for rel in self.inter_module_relations:
            type_counts[rel['relation_type']] += 1
        
        print("\n📊 Распределение по типам связей:")
        for rel_type, count in sorted(type_counts.items()):
            percentage = count / self.stats['inter_module_relations'] * 100
            print(f"   {rel_type:20} {count:4} ({percentage:5.1f}%)")
        
        # Матрица в консоли (упрощенная)
        print("\n🧮 Матрица связей (первые 5 модулей):")
        
        all_modules = sorted(set(
            list(self.connection_matrix.keys()) + 
            [t for src in self.connection_matrix.values() for t in src.keys()]
        ))[:5]
        
        if all_modules:
            print(" " * 30 + " ".join([f"{m[-10:]:>10}" for m in all_modules]))
            for source in all_modules:
                print(f"{source[:30]:<30}", end="")
                for target in all_modules:
                    count = self.connection_matrix[source].get(target, 0)
                    if count > 0:
                        print(f"{count:>10}", end="")
                    else:
                        print(" " * 10, end="")
                print()


class LinkVisitor(ast.NodeVisitor):
    """Посетитель AST для быстрого поиска связей."""
    
    def __init__(self, source_module: str, module_index: Dict[str, Path], 
                 element_cache: Dict, quick_lookup: Dict, lines: List[str]):
        self.source_module = source_module
        self.module_index = module_index
        self.element_cache = element_cache
        self.quick_lookup = quick_lookup
        self.lines = lines
        
        # Стек контекста
        self.context_stack = []
        self.relations = []
        
    def visit_ClassDef(self, node):
        """Обрабатывает определение класса."""
        self.context_stack.append(('class', node.name))
        self.generic_visit(node)
        self.context_stack.pop()
    
    def visit_FunctionDef(self, node):
        """Обрабатывает определение функции/метода."""
        # Определяем тип контекста
        context_type = 'function'
        if any(ctx[0] == 'class' for ctx in self.context_stack):
            context_type = 'method'
        
        self.context_stack.append((context_type, node.name))
        self.generic_visit(node)
        self.context_stack.pop()
    
    def visit_AsyncFunctionDef(self, node):
        """Обрабатывает асинхронную функцию/метод."""
        self.visit_FunctionDef(node)  # Та же логика
    
    def visit_Import(self, node):
        """Обрабатывает импорты."""
        context = self._get_current_context()
        
        for alias in node.names:
            imported_name = alias.name.split('.')[0]
            target_module = self._resolve_module_name(imported_name)
            
            if target_module and target_module != self.source_module:
                self.relations.append({
                    'source_module': self.source_module,
                    'source_context': context,
                    'source_line': node.lineno,
                    'source_line_text': self._get_line_text(node.lineno),
                    'target_module': target_module,
                    'target_element': '',
                    'relation_type': 'import'
                })
    
    def visit_ImportFrom(self, node):
        """Обрабатывает импорты from ... import."""
        if node.module:
            context = self._get_current_context()
            base_module = node.module.split('.')[0]
            target_module = self._resolve_module_name(base_module)
            
            if target_module and target_module != self.source_module:
                for alias in node.names:
                    self.relations.append({
                        'source_module': self.source_module,
                        'source_context': context,
                        'source_line': node.lineno,
                        'source_line_text': self._get_line_text(node.lineno),
                        'target_module': target_module,
                        'target_element': alias.name,
                        'relation_type': 'import_from'
                    })
    
    def visit_Call(self, node):
        """Обрабатывает вызовы функций."""
        context = self._get_current_context()
        
        # Пытаемся определить, что вызывается
        func_name = self._extract_name(node.func)
        if func_name:
            # Ищем элемент в кэше
            target_info = self._find_element(func_name)
            if target_info and target_info['module'] != self.source_module:
                self.relations.append({
                    'source_module': self.source_module,
                    'source_context': context,
                    'source_line': node.lineno,
                    'source_line_text': self._get_line_text(node.lineno),
                    'target_module': target_info['module'],
                    'target_element': target_info.get('name', func_name),
                    'relation_type': 'call'
                })
    
    def _get_current_context(self) -> Dict:
        """Возвращает текущий контекст (класс/функция/метод)."""
        context = {}
        
        for ctx_type, ctx_name in reversed(self.context_stack):
            if ctx_type == 'class':
                context['class'] = ctx_name
            elif ctx_type == 'method':
                context['method'] = ctx_name
            elif ctx_type == 'function':
                context['function'] = ctx_name
        
        return context
    
    def _resolve_module_name(self, module_name: str) -> Optional[str]:
        """Разрешает имя модуля."""
        if module_name in self.module_index:
            return module_name
        
        for known_module in self.module_index.keys():
            if known_module.startswith(module_name + '.') or known_module == module_name:
                return known_module
        
        return None
    
    def _extract_name(self, node: ast.AST) -> Optional[str]:
        """Извлекает имя из узла AST."""
        if isinstance(node, ast.Name):
            return node.id
        elif isinstance(node, ast.Attribute):
            return node.attr
        return None
    
    def _find_element(self, element_name: str) -> Optional[Dict]:
        """Ищет элемент в кэше."""
        # Быстрый поиск по индексу
        if element_name in self.quick_lookup:
            for element in self.quick_lookup[element_name]:
                return element
        
        return None
    
    def _get_line_text(self, line_num: int) -> str:
        """Получает текст строки."""
        if 0 < line_num <= len(self.lines):
            return self.lines[line_num - 1].strip()[:100]
        return ""


def main():
    """Основная функция."""
    parser = argparse.ArgumentParser(
        description='Генератор Excel-отчета по межмодульным зависимостям'
    )
    parser.add_argument(
        'path',
        nargs='?',
        default='.',
        help='Путь к проекту'
    )
    parser.add_argument(
        '-o', '--output',
        default='inter_module_dependencies.xlsx',
        help='Имя выходного файла'
    )
    parser.add_argument(
        '--exclude',
        nargs='+',
        default=[],
        help='Директории для исключения'
    )
    parser.add_argument(
        '--timeout',
        type=int,
        default=60,
        help='Таймаут анализа (секунды)'
    )
    
    args = parser.parse_args()
    
    if not os.path.exists(args.path):
        print(f"❌ Ошибка: путь '{args.path}' не существует")
        sys.exit(1)
    
    print(f"🚀 Запуск анализа...")
    print(f"   Проект: {args.path}")
    print(f"   Выходной файл: {args.output}")
    if args.exclude:
        print(f"   Исключаемые директории: {', '.join(args.exclude)}")
    
    try:
        reporter = OptimizedDependencyReporter(args.path, args.exclude)
        reporter.analyze_project()
        reporter.create_excel_report(args.output)
        
    except KeyboardInterrupt:
        print("\n\n⏹️  Прервано пользователем")
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()