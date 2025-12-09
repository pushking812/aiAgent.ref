#!/usr/bin/env python3
"""
Скрипт для создания Excel-отчета по связям между модулями Python.
Отчет показывает: кто вызывает (источник) и к кому обращается (цель).
"""

import os
import ast
import sys
from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Set, Tuple, Optional, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse

class DependencyReporter:
    def __init__(self, root_path: str, exclude_dirs: List[str] = None):
        """
        Инициализация репортера зависимостей.
        
        Args:
            root_path: Корневой путь проекта
            exclude_dirs: Директории для исключения
        """
        self.root_path = Path(root_path).absolute()
        self.exclude_dirs = exclude_dirs or []
        self.exclude_dirs.extend(['venv', '.venv', '__pycache__', '.git', '.idea'])
        
        # Индексы для быстрого поиска
        self.module_index: Dict[str, Path] = {}
        self.import_relations: List[Dict] = []
        self.function_calls: List[Dict] = []
        self.class_inheritance: List[Dict] = []
        
        # Статистика
        self.stats = {
            'modules': 0,
            'imports': 0,
            'calls': 0,
            'inheritance': 0,
            'files_analyzed': 0
        }
    
    def analyze_project(self) -> None:
        """Анализирует проект и собирает все связи."""
        print(f"🔍 Анализ проекта: {self.root_path}")
        
        # 1. Индексируем все модули
        self._index_modules()
        
        # 2. Анализируем каждый модуль
        for module_name, file_path in self.module_index.items():
            self._analyze_module(module_name, file_path)
        
        print(f"✅ Анализ завершен. Найдено:")
        print(f"   Модулей: {self.stats['modules']}")
        print(f"   Импортов: {self.stats['imports']}")
        print(f"   Вызовов: {self.stats['calls']}")
        print(f"   Наследований: {self.stats['inheritance']}")
    
    def _index_modules(self) -> None:
        """Индексирует все Python-модули в проекте."""
        for py_file in self.root_path.rglob("*.py"):
            # Пропускаем исключенные директории
            if any(excluded in py_file.parts for excluded in self.exclude_dirs):
                continue
            
            # Получаем относительный путь и имя модуля
            rel_path = py_file.relative_to(self.root_path)
            
            if py_file.name == '__init__.py':
                # Это пакет
                module_name = str(rel_path.parent).replace(os.sep, '.')
            else:
                # Обычный модуль
                module_name = str(rel_path.with_suffix('')).replace(os.sep, '.')
            
            self.module_index[module_name] = py_file
        
        self.stats['modules'] = len(self.module_index)
    
    def _analyze_module(self, module_name: str, file_path: Path) -> None:
        """Анализирует один модуль."""
        self.stats['files_analyzed'] += 1
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                lines = content.split('\n')
            
            tree = ast.parse(content)
            
            # Анализируем импорты
            self._analyze_imports(module_name, file_path, tree, lines)
            
            # Анализируем вызовы функций
            self._analyze_calls(module_name, file_path, tree, lines)
            
            # Анализируем наследование
            self._analyze_inheritance(module_name, file_path, tree, lines)
            
        except (SyntaxError, UnicodeDecodeError) as e:
            print(f"⚠️  Ошибка при анализе {file_path}: {e}")
    
    def _analyze_imports(self, module_name: str, file_path: Path, 
                        tree: ast.AST, lines: List[str]) -> None:
        """Анализирует импорты в модуле."""
        for node in ast.walk(tree):
            source_info = self._get_source_info(node, file_path, lines)
            
            if isinstance(node, ast.Import):
                for alias in node.names:
                    imported_name = alias.name
                    
                    # Проверяем, является ли это внутренним модулем
                    target_module = self._resolve_module_name(imported_name)
                    if target_module:
                        self.import_relations.append({
                            'source_type': 'import',
                            'source_module': module_name,
                            'source_class': source_info.get('class'),
                            'source_method': source_info.get('method'),
                            'source_line': source_info.get('line', node.lineno),
                            'source_line_text': lines[node.lineno - 1].strip() if node.lineno <= len(lines) else '',
                            'target_module': target_module,
                            'target_class': '',
                            'target_method': '',
                            'target_line': 0,
                            'target_line_text': '',
                            'relation_type': 'import'
                        })
                        self.stats['imports'] += 1
            
            elif isinstance(node, ast.ImportFrom):
                if node.module:
                    base_module = node.module
                    
                    for alias in node.names:
                        imported_name = alias.name
                        target_module = self._resolve_module_name(base_module)
                        
                        if target_module:
                            self.import_relations.append({
                                'source_type': 'import_from',
                                'source_module': module_name,
                                'source_class': source_info.get('class'),
                                'source_method': source_info.get('method'),
                                'source_line': source_info.get('line', node.lineno),
                                'source_line_text': lines[node.lineno - 1].strip() if node.lineno <= len(lines) else '',
                                'target_module': target_module,
                                'target_class': '',
                                'target_method': imported_name,
                                'target_line': 0,
                                'target_line_text': '',
                                'relation_type': 'import_from'
                            })
                            self.stats['imports'] += 1
    
    def _analyze_calls(self, module_name: str, file_path: Path, 
                      tree: ast.AST, lines: List[str]) -> None:
        """Анализирует вызовы функций."""
        for node in ast.walk(tree):
            if isinstance(node, ast.Call):
                source_info = self._get_source_info(node, file_path, lines)
                
                # Анализируем функцию, которую вызывают
                func_name = self._extract_function_name(node.func)
                if func_name:
                    # Пытаемся найти, откуда эта функция
                    target_info = self._resolve_function_call(func_name, module_name)
                    
                    if target_info:
                        self.function_calls.append({
                            'source_type': 'call',
                            'source_module': module_name,
                            'source_class': source_info.get('class'),
                            'source_method': source_info.get('method'),
                            'source_line': source_info.get('line', node.lineno),
                            'source_line_text': lines[node.lineno - 1].strip() if node.lineno <= len(lines) else '',
                            'target_module': target_info.get('module', ''),
                            'target_class': target_info.get('class', ''),
                            'target_method': target_info.get('method', func_name),
                            'target_line': target_info.get('line', 0),
                            'target_line_text': target_info.get('line_text', ''),
                            'relation_type': 'function_call'
                        })
                        self.stats['calls'] += 1
    
    def _analyze_inheritance(self, module_name: str, file_path: Path,
                           tree: ast.AST, lines: List[str]) -> None:
        """Анализирует отношения наследования."""
        for node in ast.walk(tree):
            if isinstance(node, ast.ClassDef):
                source_info = self._get_source_info(node, file_path, lines)
                
                # Анализируем базовые классы
                for base in node.bases:
                    if isinstance(base, ast.Name):
                        base_name = base.id
                        
                        # Пытаемся найти класс в проекте
                        target_info = self._resolve_class_reference(base_name, module_name)
                        
                        if target_info:
                            self.class_inheritance.append({
                                'source_type': 'inheritance',
                                'source_module': module_name,
                                'source_class': node.name,
                                'source_method': '',
                                'source_line': source_info.get('line', node.lineno),
                                'source_line_text': lines[node.lineno - 1].strip() if node.lineno <= len(lines) else '',
                                'target_module': target_info.get('module', ''),
                                'target_class': target_info.get('class', base_name),
                                'target_method': '',
                                'target_line': target_info.get('line', 0),
                                'target_line_text': target_info.get('line_text', ''),
                                'relation_type': 'inheritance'
                            })
                            self.stats['inheritance'] += 1
    
    def _get_source_info(self, node: ast.AST, file_path: Path, 
                        lines: List[str]) -> Dict[str, Any]:
        """Получает информацию об источнике вызова."""
        info = {'line': node.lineno if hasattr(node, 'lineno') else 0}
        
        # Находим родительские узлы для определения контекста
        parent = getattr(node, 'parent', None)
        
        # Ищем класс и метод в родительских узлах
        current = node
        while hasattr(current, 'parent'):
            current = current.parent
            if isinstance(current, ast.ClassDef):
                info['class'] = current.name
                break
        
        # Сбрасываем для поиска метода
        current = node
        while hasattr(current, 'parent'):
            current = current.parent
            if isinstance(current, ast.FunctionDef):
                info['method'] = current.name
                break
        
        return info
    
    def _resolve_module_name(self, module_name: str) -> Optional[str]:
        """Разрешает имя модуля до полного имени в проекте."""
        # Убираем возможные относительные импорты
        if module_name.startswith('.'):
            # Пропускаем относительные импорты для упрощения
            return None
        
        # Проверяем точное совпадение
        if module_name in self.module_index:
            return module_name
        
        # Проверяем частичное совпадение (модуль может быть частью пути)
        for known_module in self.module_index.keys():
            if known_module == module_name or known_module.startswith(module_name + '.'):
                return known_module
        
        return None
    
    def _extract_function_name(self, func_node: ast.AST) -> Optional[str]:
        """Извлекает имя функции из узла AST."""
        if isinstance(func_node, ast.Name):
            return func_node.id
        elif isinstance(func_node, ast.Attribute):
            # Рекурсивно извлекаем цепочку атрибутов
            parts = []
            current = func_node
            
            while isinstance(current, ast.Attribute):
                parts.append(current.attr)
                current = current.value
            
            if isinstance(current, ast.Name):
                parts.append(current.id)
            
            return '.'.join(reversed(parts))
        
        return None
    
    def _resolve_function_call(self, func_name: str, 
                             source_module: str) -> Optional[Dict]:
        """Пытается разрешить вызов функции до ее определения."""
        # Упрощенная реализация - ищем в индексированных модулях
        # В реальном проекте нужно более сложное разрешение
        
        # Разделяем имя функции на части
        parts = func_name.split('.')
        
        if len(parts) == 1:
            # Простое имя функции - ищем в том же модуле
            return self._find_function_in_module(parts[0], source_module)
        elif len(parts) > 1:
            # Возможно, это вызов метода или функция из другого модуля
            module_part = '.'.join(parts[:-1])
            func_part = parts[-1]
            
            # Сначала проверяем, не является ли это модулем
            target_module = self._resolve_module_name(module_part)
            if target_module:
                return self._find_function_in_module(func_part, target_module)
        
        return None
    
    def _find_function_in_module(self, func_name: str, 
                               module_name: str) -> Optional[Dict]:
        """Ищет функцию в указанном модуле."""
        if module_name not in self.module_index:
            return None
        
        file_path = self.module_index[module_name]
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                lines = content.split('\n')
            
            tree = ast.parse(content)
            
            for node in ast.walk(tree):
                if isinstance(node, ast.FunctionDef) and node.name == func_name:
                    return {
                        'module': module_name,
                        'method': func_name,
                        'line': node.lineno,
                        'line_text': lines[node.lineno - 1].strip() if node.lineno <= len(lines) else ''
                    }
                
                elif isinstance(node, ast.ClassDef):
                    # Ищем метод в классе
                    for item in node.body:
                        if isinstance(item, ast.FunctionDef) and item.name == func_name:
                            return {
                                'module': module_name,
                                'class': node.name,
                                'method': func_name,
                                'line': item.lineno,
                                'line_text': lines[item.lineno - 1].strip() if item.lineno <= len(lines) else ''
                            }
        
        except Exception:
            pass
        
        return None
    
    def _resolve_class_reference(self, class_name: str, 
                               source_module: str) -> Optional[Dict]:
        """Разрешает ссылку на класс."""
        # Ищем класс в модулях проекта
        for module_name, file_path in self.module_index.items():
            if module_name == source_module:
                # Пропускаем тот же модуль (для простоты)
                continue
            
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    lines = content.split('\n')
                
                tree = ast.parse(content)
                
                for node in ast.walk(tree):
                    if isinstance(node, ast.ClassDef) and node.name == class_name:
                        return {
                            'module': module_name,
                            'class': class_name,
                            'line': node.lineno,
                            'line_text': lines[node.lineno - 1].strip() if node.lineno <= len(lines) else ''
                        }
            
            except Exception:
                continue
        
        return None
    
    def create_excel_report(self, output_file: str = "module_dependencies.xlsx") -> None:
        """Создает Excel-отчет с зависимостями."""
        print(f"\n📊 Создание Excel-отчета...")
        
        # Объединяем все связи
        all_relations = []
        all_relations.extend(self.import_relations)
        all_relations.extend(self.function_calls)
        all_relations.extend(self.class_inheritance)
        
        if not all_relations:
            print("⚠️  Не найдено зависимостей для отчета")
            return
        
        # Создаем DataFrame
        df = pd.DataFrame(all_relations)
        
        # Сортируем по типу связи
        if 'relation_type' in df.columns:
            df = df.sort_values(['relation_type', 'source_module', 'source_line'])
        
        # Создаем Excel-файл с форматированием
        self._create_formatted_excel(df, output_file)
        
        print(f"✅ Отчет сохранен: {output_file}")
        
        # Выводим сводку
        self._print_summary(df)
    
    def _create_formatted_excel(self, df: pd.DataFrame, output_file: str) -> None:
        """Создает форматированный Excel-файл."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Зависимости модулей"
        
        # Заголовки столбцов
        headers = [
            # Источник (кто вызывает)
            'Тип связи',
            'Источник: Папка',
            'Источник: Модуль', 
            'Источник: Класс',
            'Источник: Метод',
            'Источник: Строка',
            'Источник: Код строки',
            
            # Цель (к кому обращаются)
            'Цель: Папка',
            'Цель: Модуль',
            'Цель: Класс',
            'Цель: Метод',
            'Цель: Строка',
            'Цель: Код строки'
        ]
        
        # Добавляем заголовки
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Добавляем данные
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            # Определяем цвет строки в зависимости от типа связи
            fill_color = self._get_row_color(row.get('relation_type', ''))
            
            # Заполняем столбцы источника
            ws.cell(row=row_idx, column=1, value=row.get('relation_type', ''))
            ws.cell(row=row_idx, column=2, value=self._extract_folder(row.get('source_module', '')))
            ws.cell(row=row_idx, column=3, value=row.get('source_module', ''))
            ws.cell(row=row_idx, column=4, value=row.get('source_class', ''))
            ws.cell(row=row_idx, column=5, value=row.get('source_method', ''))
            ws.cell(row=row_idx, column=6, value=row.get('source_line', 0))
            ws.cell(row=row_idx, column=7, value=row.get('source_line_text', ''))
            
            # Заполняем столбцы цели
            ws.cell(row=row_idx, column=8, value=self._extract_folder(row.get('target_module', '')))
            ws.cell(row=row_idx, column=9, value=row.get('target_module', ''))
            ws.cell(row=row_idx, column=10, value=row.get('target_class', ''))
            ws.cell(row=row_idx, column=11, value=row.get('target_method', ''))
            ws.cell(row=row_idx, column=12, value=row.get('target_line', 0))
            ws.cell(row=row_idx, column=13, value=row.get('target_line_text', ''))
            
            # Применяем форматирование ко всей строке
            for col_idx in range(1, 14):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                # Добавляем границы
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # Настраиваем ширину столбцов
        column_widths = {
            'A': 15,  # Тип связи
            'B': 20,  # Папка источника
            'C': 30,  # Модуль источника
            'D': 20,  # Класс источника
            'E': 20,  # Метод источника
            'F': 10,  # Строка источника
            'G': 40,  # Код источника
            'H': 20,  # Папка цели
            'I': 30,  # Модуль цели
            'J': 20,  # Класс цели
            'K': 20,  # Метод цели
            'L': 10,  # Строка цели
            'M': 40   # Код цели
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Замораживаем заголовки
        ws.freeze_panes = 'A2'
        
        # Создаем лист со сводкой
        self._create_summary_sheet(wb)
        
        # Сохраняем файл
        wb.save(output_file)
    
    def _extract_folder(self, module_name: str) -> str:
        """Извлекает папку из имени модуля."""
        if not module_name:
            return ''
        
        parts = module_name.split('.')
        if len(parts) > 1:
            return '.'.join(parts[:-1])
        return module_name
    
    def _get_row_color(self, relation_type: str) -> str:
        """Возвращает цвет строки в зависимости от типа связи."""
        colors = {
            'import': 'E6F3FF',  # Светло-голубой
            'import_from': 'D9F2E6',  # Светло-зеленый
            'function_call': 'FFF2E6',  # Светло-оранжевый
            'inheritance': 'F2E6FF',  # Светло-фиолетовый
        }
        return colors.get(relation_type, 'FFFFFF')
    
    def _create_summary_sheet(self, wb: Workbook) -> None:
        """Создает лист со сводкой."""
        ws_summary = wb.create_sheet(title="Сводка")
        
        # Заголовок
        ws_summary.merge_cells('A1:D1')
        title_cell = ws_summary.cell(row=1, column=1, 
                                   value="СВОДКА ЗАВИСИМОСТЕЙ МОДУЛЕЙ")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Информация о проекте
        ws_summary.cell(row=3, column=1, value="Проект:").font = Font(bold=True)
        ws_summary.cell(row=3, column=2, value=str(self.root_path))
        
        ws_summary.cell(row=4, column=1, value="Дата анализа:").font = Font(bold=True)
        from datetime import datetime
        ws_summary.cell(row=4, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Статистика
        row = 6
        ws_summary.cell(row=row, column=1, value="СТАТИСТИКА").font = Font(bold=True, size=12)
        
        stats_data = [
            ("Модулей проанализировано", self.stats['files_analyzed']),
            ("Всего модулей в проекте", self.stats['modules']),
            ("Импортов найдено", self.stats['imports']),
            ("Вызовов функций", self.stats['calls']),
            ("Отношений наследования", self.stats['inheritance']),
            ("Всего зависимостей", 
             self.stats['imports'] + self.stats['calls'] + self.stats['inheritance'])
        ]
        
        for i, (label, value) in enumerate(stats_data, start=row+1):
            ws_summary.cell(row=i, column=1, value=label)
            ws_summary.cell(row=i, column=2, value=value)
        
        # Самые используемые модули
        row = len(stats_data) + row + 2
        ws_summary.cell(row=row, column=1, value="САМЫЕ ИСПОЛЬЗУЕМЫЕ МОДУЛИ").font = Font(bold=True, size=12)
        
        # Анализируем статистику использования
        usage_stats = defaultdict(int)
        for rel in self.import_relations + self.function_calls + self.class_inheritance:
            target = rel.get('target_module', '')
            if target:
                usage_stats[target] += 1
        
        sorted_usage = sorted(usage_stats.items(), key=lambda x: x[1], reverse=True)
        
        for i, (module, count) in enumerate(sorted_usage[:10], start=row+1):
            ws_summary.cell(row=i, column=1, value=module)
            ws_summary.cell(row=i, column=2, value=count)
        
        # Настраиваем ширину столбцов
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
    
    def _print_summary(self, df: pd.DataFrame) -> None:
        """Выводит сводку в консоль."""
        print(f"\n📋 СВОДКА ОТЧЕТА:")
        print(f"{'='*60}")
        
        # Группируем по типу связи
        if 'relation_type' in df.columns:
            type_counts = df['relation_type'].value_counts()
            print("\n📈 Распределение по типам связей:")
            for relation_type, count in type_counts.items():
                print(f"   {relation_type}: {count}")
        
        # Топ-5 модулей по количеству зависимостей
        if 'source_module' in df.columns:
            source_counts = df['source_module'].value_counts().head(5)
            print("\n🏆 Топ-5 модулей с наибольшим количеством исходящих связей:")
            for module, count in source_counts.items():
                print(f"   {module}: {count} связей")
        
        if 'target_module' in df.columns:
            target_counts = df['target_module'].value_counts().head(5)
            print("\n🎯 Топ-5 наиболее используемых модулей:")
            for module, count in target_counts.items():
                print(f"   {module}: {count} ссылок")
        
        print(f"\n📄 Полный отчет доступен в Excel-файле")

def main():
    """Основная функция."""
    parser = argparse.ArgumentParser(
        description='Генератор Excel-отчета по зависимостям Python-модулей'
    )
    parser.add_argument(
        'path',
        nargs='?',
        default='.',
        help='Путь к проекту (по умолчанию: текущая директория)'
    )
    parser.add_argument(
        '-o', '--output',
        default='module_dependencies.xlsx',
        help='Имя выходного Excel-файла (по умолчанию: module_dependencies.xlsx)'
    )
    parser.add_argument(
        '--exclude',
        nargs='+',
        default=[],
        help='Дополнительные директории для исключения'
    )
    parser.add_argument(
        '--simple',
        action='store_true',
        help='Простой анализ (только импорты)'
    )
    
    args = parser.parse_args()
    
    # Проверяем существование пути
    if not os.path.exists(args.path):
        print(f"❌ Ошибка: путь '{args.path}' не существует")
        sys.exit(1)
    
    # Создаем репортер
    reporter = DependencyReporter(args.path, exclude_dirs=args.exclude)
    
    try:
        # Анализируем проект
        reporter.analyze_project()
        
        # Создаем отчет
        reporter.create_excel_report(args.output)
        
    except KeyboardInterrupt:
        print("\n\n⏹️  Прервано пользователем")
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()